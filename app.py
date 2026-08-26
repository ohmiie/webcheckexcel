import streamlit as st
import openpyxl
import pandas as pd

# ---------------------------------------------------------
# การตั้งค่าหน้าเว็บ
# ---------------------------------------------------------
st.set_page_config(page_title="ระบบประเมินทักษะ Excel", layout="wide", page_icon="📊")

st.title("📊 ระบบประเมินและให้คะแนนทักษะวิชาการ Excel")
st.markdown("อัปโหลดไฟล์เพื่อตรวจคะแนนทันที ระบบจะไม่บันทึกหรือจดจำข้อมูลค้างไว้เพื่อความรวดเร็ว")

# ==========================================
# อัปโหลดไฟล์ (แบบเพียวๆ ไม่มีดึงชื่อ)
# ==========================================
uploaded_file = st.file_uploader("📂 อัปโหลดไฟล์ Excel (.xlsx)", type=["xlsx"])

# ตัวแปรเก็บคะแนนแต่ละหมวด
c1, c2, c3, c4, c5, c6, c7, c8, c9 = 0, 0, 0, 0, 0, 0, 0, 0, 0

def get_safe_formula(cell):
    if cell and cell.value:
        return str(cell.value).upper()
    return ""

if uploaded_file is not None:
    st.info("🔄 กำลังประมวลผลไฟล์ กรุณารอสักครู่...")
    
    try:
        # โหลดไฟล์
        wb_formula = openpyxl.load_workbook(uploaded_file, data_only=False)
        wb_data = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet_names = wb_formula.sheetnames
        
        col_left, col_right = st.columns([1.5, 1])
        
        with col_left:
            st.header("🤖 ส่วนตรวจอัตโนมัติ (Auto-Grading)")
            
            # --- หมวด 1: เตรียมแผ่นงาน (1 คะแนน) ---
            st.markdown("#### 1. การจัดเตรียมแผ่นงาน")
            if "Student_Scores" in sheet_names:
                c1 += 1
                st.success("✅ เปลี่ยนชื่อ Sheet เป็น 'Student_Scores' (ได้ 1 คะแนน)")
            else:
                st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบชีต 'Student_Scores' (คะแนน 0)")
            
            # --- หมวด 2: คำนวณ Student_Scores (สูตร 18 + Manual 6 = 24 คะแนน) ---
            st.markdown("#### 2. การคำนวณข้อมูลใน Sheet 'Student_Scores'")
            if "Student_Scores" in sheet_names:
                ws_form = wb_formula["Student_Scores"]
                ws_dat = wb_data["Student_Scores"]
                
                if ws_dat.max_row >= 50: 
                    c2 += 2
                    st.success("✅ ป้อนข้อมูลนักศึกษาครบถ้วน (ได้ 2 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ข้อมูลไม่ครบ 50 รายการ (คะแนน 0)")
                    
                if ws_dat.max_column >= 10: 
                    c2 += 2
                    st.success("✅ นำเข้าข้อมูลคะแนนดิบและอื่นๆ ครบถ้วน (ได้ 2 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ข้อมูลคอลัมน์ไม่ครบถ้วน (คะแนน 0)")
                
                if any("SUM(" in get_safe_formula(ws_form[f'I{r}']) for r in range(3, 10)): 
                    c2 += 3
                    st.success("✅ คำนวณผลรวมคะแนนดิบด้วยฟังก์ชัน SUM (ได้ 3 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบสูตร SUM ในคอลัมน์ผลรวม (คะแนน 0)")
                    
                if any("IF(" in get_safe_formula(ws_form[f'K{r}']) for r in range(3, 10)): 
                    c2 += 4
                    st.success("✅ คำนวณคะแนนหักขาดเรียนด้วย IF (ได้ 4 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบสูตร IF ในการหักคะแนนขาดเรียน (คะแนน 0)")
                    
                if any("-" in get_safe_formula(ws_form[f'L{r}']) for r in range(3, 10)): 
                    c2 += 2
                    st.success("✅ คำนวณคะแนนสุทธิถูกต้อง (ได้ 2 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบการคำนวณคะแนนสุทธิ (คะแนน 0)")
                
                if "Net_Score_Data" in wb_formula.defined_names: 
                    c2 += 2
                    st.success("✅ กำหนดชื่อ Named Range 'Net_Score_Data' ถูกต้อง (ได้ 2 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบ Named Range ชื่อ 'Net_Score_Data' (คะแนน 0)")
                
                if any("RANK" in get_safe_formula(ws_form[f'M{r}']) for r in range(3, 10)): 
                    c2 += 3
                    st.success("✅ หาอันดับด้วยฟังก์ชัน RANK หรือ RANK.EQ (ได้ 3 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบสูตร RANK หาอันดับคะแนน (คะแนน 0)")
            else:
                st.error("❌ ข้ามการตรวจหมวด 2 เนื่องจากไม่พบชีต 'Student_Scores'")

            # --- หมวด 3: Student_Scores2 (สูตร 4 + Manual 3 = 7 คะแนน) ---
            st.markdown("#### 3. การเชื่อมโยงชีต 'Student_Scores2'")
            if "Student_Scores2" in sheet_names:
                c3 += 1
                st.success("✅ เปลี่ยนชื่อ Sheet เป็น 'Student_Scores2' (ได้ 1 คะแนน)")
                ws_form2 = wb_formula["Student_Scores2"]
                if any("STUDENT_SCORES!" in get_safe_formula(ws_form2[f'B{r}']) for r in range(2, 10)): 
                    c3 += 3
                    st.success("✅ ดึงข้อมูลเชื่อมโยง (Link Sheet) มาวางถูกต้อง (ได้ 3 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบการทำ Link Sheet (คะแนน 0)")
            else:
                st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบชีต 'Student_Scores2' (คะแนน 0)")

            # --- หมวด 5: สร้าง Grade_Summary (สูตร 1 + Manual 1 = 2 คะแนน) ---
            st.markdown("#### 5. ชีต 'Grade_Summary'")
            if "Grade_Summary" in sheet_names:
                c5 += 1
                st.success("✅ สร้างและเปลี่ยนชื่อ Sheet เป็น 'Grade_Summary' (ได้ 1 คะแนน)")
            else:
                st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบชีต 'Grade_Summary' (คะแนน 0)")

            # --- หมวด 6: คำนวณ Grade_Summary (สูตร 18 = 18 คะแนน) ---
            st.markdown("#### 6. การคำนวณเกรด 'Grade_Summary'")
            if "Grade_Summary" in sheet_names:
                ws_grade = wb_formula["Grade_Summary"]
                if any("STUDENT_SCORES!" in get_safe_formula(ws_grade[f'B{r}']) for r in range(3, 10)): 
                    c6 += 4
                    st.success("✅ ดึงข้อมูลเชื่อมโยงมาถูกต้อง (ได้ 4 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบการ Link Sheet ในตารางเกรด (คะแนน 0)")
                    
                if any(get_safe_formula(ws_grade[f'F{r}']).count("IF(") >= 4 for r in range(3, 10)): 
                    c6 += 8
                    st.success("✅ ใช้ Nested IF ตัดเกรด A,B,C,D,F ถูกต้อง (ได้ 8 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบการใช้ Nested IF หรือซ้อนกันไม่ครบ (คะแนน 0)")
                    
                if any("IF(" in get_safe_formula(ws_grade[f'G{r}']) for r in range(3, 10)): 
                    c6 += 6
                    st.success("✅ ใช้ฟังก์ชัน IF กำหนดสถานะผ่าน/ไม่ผ่าน (ได้ 6 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบฟังก์ชัน IF กำหนดสถานะ (คะแนน 0)")
            else:
                st.error("❌ ข้ามการตรวจหมวด 6 เนื่องจากไม่พบชีต 'Grade_Summary'")

            # --- หมวด 7: แดชบอร์ด (การเพิ่ม Sheet 2 คะแนน) ---
            st.markdown("#### 7. ชีต 'Report_Dashboard'")
            if "Report_Dashboard" in sheet_names:
                c7 += 2
                st.success("✅ สร้างและเปลี่ยนชื่อ Sheet เป็น 'Report_Dashboard' (ได้ 2 คะแนน)")
            else:
                st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบชีต 'Report_Dashboard' (คะแนน 0)")
            
            # --- หมวด 8: แดชบอร์ดสรุปผล (สูตร 12 + Manual 12 = 24 คะแนน) ---
            st.markdown("#### 8. การคำนวณแดชบอร์ดสรุปผล")
            if "Report_Dashboard" in sheet_names:
                ws_dash = wb_formula["Report_Dashboard"]
                dash_formulas = " ".join([get_safe_formula(cell) for row in ws_dash.iter_rows() for cell in row if cell.data_type == 'f'])
                
                if "COUNT(" in dash_formulas or "COUNTA(" in dash_formulas: 
                    c8 += 2
                    st.success("✅ ใช้ฟังก์ชัน COUNTA/COUNT หานักศึกษาทั้งหมด (ได้ 2 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบฟังก์ชัน COUNT/COUNTA (คะแนน 0)")
                    
                if dash_formulas.count("COUNTIF(") >= 2: 
                    c8 += 10
                    st.success("✅ ใช้ฟังก์ชัน COUNTIF นับเกรด/กลุ่ม/สถานะ (ได้ 10 คะแนน)")
                else: 
                    st.error("❌ ไม่ตรงกับเกณฑ์: ไม่พบการใช้ COUNTIF ที่เพียงพอ (คะแนน 0)")
            else:
                st.error("❌ ข้ามการตรวจหมวด 8 เนื่องจากไม่พบชีต 'Report_Dashboard'")

        with col_right:
            st.header("✍️ ส่วนตรวจด้วยสายตา (Manual)")
            
            with st.expander("คะแนนส่วนกราฟและรูปแบบ", expanded=True):
                if st.checkbox("2.7 แทรกกราฟ Line Sparklines (3 คะแนน)"): c2 += 3
                if st.checkbox("2.8 คำนวณ Average, Max, Min ท้ายตาราง (3 คะแนน)"): c2 += 3
                if st.checkbox("3.3 จัดเรียงข้อมูล (Sort) 1-50 (3 คะแนน)"): c3 += 3
                if st.checkbox("5.2 โครงสร้างหัวตาราง Grade_Summary (1 คะแนน)"): c5 += 1
                if st.checkbox("7.2 โครงสร้าง Dashboard สวยงาม (3 คะแนน)"): c8 += 3
            
            with st.expander("ข้อ 4: Report_Table (12 คะแนน)", expanded=True):
                if st.checkbox("4.1 มีชีต 'Report_Table' (1 คะแนน)"): c4 += 1
                if st.checkbox("4.2 สร้าง PivotTable สรุปถูกต้อง (5 คะแนน)"): c4 += 5
                if st.checkbox("4.3 สร้าง PivotChart รูปแบบ Column (3 คะแนน)"): c4 += 3
                if st.checkbox("4.4 มีเครื่องมือ Slicer 'กลุ่มเรียน' (3 คะแนน)"): c4 += 3
                
            with st.expander("ข้อ 7: Report_Dashboard", expanded=True):
                if st.checkbox("7.7 สร้าง PivotTable สรุปเกรด (4 คะแนน)"): c8 += 4
                if st.checkbox("7.8 สร้าง PivotChart รูปแบบ Column (3 คะแนน)"): c8 += 3
                if st.checkbox("7.9 มีเครื่องมือ Slicer 'กลุ่มเรียน' (2 คะแนน)"): c8 += 2

            with st.expander("ข้อ 8: ไฟล์ PDF และการพิมพ์ (10 คะแนน)", expanded=True):
                if st.checkbox("8.1 Grade_Summary แนวตั้ง A4, Scale 90% (2 คะแนน)"): c9 += 2
                if st.checkbox("8.1.1 Margins (บน/ล่าง 1, ซ้าย/ขวา 0.5, H/F 1) (2 คะแนน)"): c9 += 2
                if st.checkbox("8.1.2 มี Header ขวา / Footer กลาง และส่งเป็น PDF (2 คะแนน)"): c9 += 2
                if st.checkbox("8.2 กราฟ SEC04 แนวนอน, A4, Margins (2 คะแนน)"): c9 += 2
                if st.checkbox("8.3 ส่งเป็น PDF ชื่อไฟล์ 'SEC04' (2 คะแนน)"): c9 += 2

        # รวมคะแนนทั้งหมด
        total_score = c1 + c2 + c3 + c4 + c5 + c6 + c7 + c8 + c9
        
        st.divider()
        st.markdown(f"<h1 style='text-align: center; color: #1f77b4;'>🏆 คะแนนรวมสุทธิ: {total_score} / 100</h1>", unsafe_allow_html=True)
        
        # ==========================================
        # ตาราง HTML CSS ตะแคง 45 องศา
        # ==========================================
        st.markdown("### 📋 สรุปคะแนน (คลุมดำแล้ว Copy ไปวางใน Excel ได้เลย)")
        
        html_table = f"""
        <style>
        .score-table-container {{
            overflow-x: auto;
            padding-top: 20px;
            padding-bottom: 20px;
            margin-bottom: 50px;
        }}
        table.custom-score-table {{
            border-collapse: collapse;
            width: 100%;
            background-color: white;
            font-family: sans-serif;
        }}
        table.custom-score-table th {{
            height: 280px;
            width: 60px;
            position: relative;
            border: 1px solid #ddd;
            background-color: #f1f3f6;
            vertical-align: bottom;
        }}
        table.custom-score-table th > div {{
            position: absolute;
            bottom: 20px;
            left: 50%;
            transform-origin: left bottom;
            transform: translate(-50%, 0) rotate(-45deg);
            white-space: nowrap;
            font-size: 13px;
            color: #333;
            line-height: 1.2;
        }}
        table.custom-score-table td {{
            border: 1px solid #ddd;
            text-align: center;
            padding: 15px;
            font-size: 18px;
            font-weight: bold;
            color: #2c3e50;
        }}
        </style>

        <div class="score-table-container">
            <table class="custom-score-table">
                <thead>
                    <tr>
                        <th><div>การจัดเตรียมแผ่นงานและบันทึกข้อมูล<br>(1 คะแนน)</div></th>
                        <th><div>การคำนวณและประมวลผลข้อมูลใน Sheet "Student_Scores"<br>(24 คะแนน)</div></th>
                        <th><div>การเชื่อมโยงและจัดเรียงข้อมูล<br>ใน Sheet "Student_Scores2" (7 คะแนน)</div></th>
                        <th><div>การรายงานตารางสรุปข้อมูลใน Sheet "Report_Table"<br>(12 คะแนน)</div></th>
                        <th><div>การสร้างแผ่นงานประมวลผลเกรด Sheet "Grade_Summary"<br>(2 คะแนน)</div></th>
                        <th><div>การคำนวณเกรดและสถานะประเมินใน Sheet "Grade_Summary"<br>(18 คะแนน)</div></th>
                        <th><div>การเพิ่มแผ่นงานแดชบอร์ด Sheet "Report_Dashboard"<br>(2 คะแนน)</div></th>
                        <th><div>การสร้างแดชบอร์ดสรุปผลและแผนภูมิใน Sheet "Report_Dashboard"<br>(24 คะแนน)</div></th>
                        <th><div>การตั้งค่าหน้ากระดาษและการพิมพ์รูปแบบ PDF<br>(10 คะแนน)</div></th>
                        <th><div>รวม<br>(100 คะแนน)</div></th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td>{c1}</td>
                        <td>{c2}</td>
                        <td>{c3}</td>
                        <td>{c4}</td>
                        <td>{c5}</td>
                        <td>{c6}</td>
                        <td>{c7}</td>
                        <td>{c8}</td>
                        <td>{c9}</td>
                        <td style="color: #e74c3c; background-color: #fdf5f6;">{total_score}</td>
                    </tr>
                </tbody>
            </table>
        </div>
        """
        
        # เรนเดอร์ HTML ออกมาที่หน้าเว็บ
        st.markdown(html_table, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"❌ เกิดข้อผิดพลาดในการโหลดไฟล์ รายละเอียด: {e}")
